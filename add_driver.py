from app import *
from flask_sqlalchemy import SQLAlchemy
from flask import Flask, flash, redirect, render_template, request, session, abort, url_for, flash
from openpyxl import load_workbook

def selection():
    choice = input('''Would you like to manually add one or more drivers,\
or would you like to scrape the Excel file? (Enter 1 or 2)
1. Manually
2. Dynamically\n''')

    if choice == "1":
        return True

def db_add(d_first, l_name, cell_num):
    new_driver = Drivers(first_name=d_first,
                        last_name=l_name,
                        c_num=cell_num,
                        )
    db.session.add(new_driver)
    db.session.commit()
    print("Added {} to list of drivers\n*****\n".format(d_first))

# Inputs for new driver
def man_add():
    d_first = input('First name:\n')
    l_name = input('Last name:\n')
    cell_num = input('Number:\n')
    db_add(d_first, l_name, cell_num)


def dyn_add():
    # Add in drivers from Excel file
    # File dependencies
    wb = load_workbook(filename='drivers.xlsx')
    ws = wb.active
    r = 1
    for row in ws.iter_rows():
        for cell in row:
            ws_row = cell.row
            # Get the driver's first name:
            if ws["D{}".format(cell.row)].value != 'x':
                d_first = ws["A{}".format(cell.row)].value
                # Get the driver's last name:
                l_name = ws["B{}".format(cell.row)].value
                # Get the driver's cell num:
                cell_num = "+1" + str(ws["C{}".format(cell.row)].value)
                # Mark that they've been added
                ws.cell(row=r, column=4).value = 'x'
                wb.save("drivers.xlsx")
            else:
                pass
        r += 1
        try:
            db_add(d_first, l_name, cell_num)
        except Exception:
            print('record already in db')


def main():
    if selection():
        while True:
            man_add()
    else:
        dyn_add()


main()
