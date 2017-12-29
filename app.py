from twilio import twiml
from flask import Flask, flash, redirect, render_template, request, session, abort, url_for, flash
from flask_wtf import FlaskForm
from flask_bootstrap import Bootstrap
from flask_sqlalchemy import SQLAlchemy
import time
import random
from random import randint
from reg_questions import questions
import openpyxl
from openpyxl import *

'''
This application allows for our school to keep track of bus ("cheese") trips,
which faculty member is driving, and who is on the bus/return trip.
'''
app = Flask(__name__)
Bootstrap(app)
app.secret_key = "space1sth3plac3"


'''
Db information
'''
# DB Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db.db'
db = SQLAlchemy(app)

# Users table for admins
class Users( db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String)
    pwd = db.Column(db.String)
    first_name = db.Column(db.String)
    last_name = db.Column(db.String)
    c_num = db.Column(db.String)

# Drivers table
class Drivers(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String)
    last_name = db.Column(db.String)
    c_num = db.Column(db.String)

# Students table
class Students(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String)
    last_name = db.Column(db.String)
    c_num = db.Column(db.String)
    m_name = db.Column(db.String)
    m_num = db.Column(db.String)
    f_name = db.Column(db.String)
    f_num = db.Column(db.String)

# Trips table
class Trips(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    d_id = db.Column(db.Integer, db.ForeignKey('drivers.id'))
    driver = db.Column(db.String, db.ForeignKey('drivers.first_name'))
    dest = db.Column(db.String)
    pin = db.Column(db.String)

    def trip_info(self):
        return('''A new trip was created by {}. The trip ID is {} and it's
               destination is {}.
               '''.format(self.driver, self.pin, self.dest))

# Passengers table
class Passengers(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    trip_pin = db.Column(db.String, db.ForeignKey('trips.pin'))
    s_id = db.Column(db.Integer, db.ForeignKey('students.id'))
    s_first = db.Column(db.Integer, db.ForeignKey('students.first_name'))
    s_last = db.Column(db.Integer, db.ForeignKey('students.last_name'))
    s_m_name = db.Column(db.String, db.ForeignKey('students.m_name'))
    s_f_name = db.Column(db.String, db.ForeignKey('students.f_name'))
    s_m_num = db.Column(db.String, db.ForeignKey('students.m_num'))
    s_f_num = db.Column(db.String, db.ForeignKey('students.f_num'))
    return_trip = db.Column(db.Boolean, default=True)

    def pass_info(self):
        return ("{} is a passenger on trip {}".format(self.s_first, self.trip_pin))

    def return_info(self):
        if self.return_trip != 1:
            return ("{} is not riding back".format(self.s_first))

    def all_info(self):
        return ('''Contact info for {} {}\n{}: {}\n{}: {}
               '''.format(self.s_first, self.s_last, self.s_m_name, self.s_m_num, self.s_f_name, self.s_f_num))



'''
Dummy Classes for user roles to validate input
'''
class Drive(object):
    def role(self, truth):
        self.truth = 1

class Student(object):
    def role(self, truth):
        self.truth = 1

class Admin(object):
    def role(self, truth):
        self.truth = 1


'''
Route handling
'''
# Index for admins
@app.route('/', methods=['GET'])
def index():

    return render_template('index.html', **locals())

# Login
@app.route('/login', methods=['GET', 'POST'])
def login():
    print('hello')

# Trips
@app.route('/trips', methods=['GET', 'POST'])
def trips():
    trip_list = Trips.query.all()
    current_trips = Trips.query.all()  # TODO: Add date attribute for a trip


    return render_template('trips.html', **locals())

# Trip detail
@app.route('/trips/<int:trip_id>')
def trip_detail(trip_id):
    trip_id = str(trip_id)
    trip = Trips.query.filter_by(pin = trip_id).first()
    manifest = Passengers.query.filter_by(trip_pin=trip_id).all()
    count = 1
    for x in manifest:
        x.pass_info()
        x.return_info()

    return render_template('trip_detail.html', **locals())

# Trip report
@app.route('/trips/export/<int:trip_id>')
def export_trip(trip_id):
    trip_id = str(trip_id)
    trip = Trips.query.filter_by(pin = trip_id).first()
    manifest = Passengers.query.filter_by(trip_pin=trip_id).all()
    # Init and set information for manifest
    filename = "trip_{}.xlsx".format(trip_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    r = 1
    ws.cell(row=r, column=1).value = "Manifest for:"
    ws.cell(row=r, column=2).value = trip_id
    r += 2
    ws.cell(row=r, column=1).value = "First Name"
    ws.cell(row=r, column=2).value = "Last Name"
    ws.cell(row=r, column=3).value = "Primary Contact"
    ws.cell(row=r, column=4).value = "Primary Number"
    ws.cell(row=r, column=5).value = "Secondary Contact"
    ws.cell(row=r, column=6).value = "Secondary Number"
    wb.save(filename)
    r += 2

    # Add students
    for x in manifest:
        for row in ws.iter_rows():
            ws.cell(row=r, column=1).value = x.s_first
            ws.cell(row=r, column=2).value = x.s_last
            ws.cell(row=r, column=3).value = x.s_m_name
            ws.cell(row=r, column=4).value = x.s_m_num
            ws.cell(row=r, column=5).value = x.s_f_name
            ws.cell(row=r, column=6).value = x.s_f_num
            wb.save(filename)


    return render_template('trip_detail.html', **locals())

'''SMS Handling via endpoint'''
@app.route('/sms', methods=['POST'])
def rec_sms():
    msg_number = request.form['From']
    message_body = request.form['Body']
    global message_body_string
    message_body_string = str(message_body)
    resp_update = twiml.Response()
    driver_numbers = Drivers.query.filter_by(c_num = msg_number).all()
    student_numbers = Students.query.filter_by(c_num = msg_number).all()
    admin_numbers = Users.query.filter_by(c_num = msg_number).all()
    reg_question_num = 0

    # Determine which type of user the msg sender is
    if len(driver_numbers) > 0:
        x = Drivers.query.filter_by(c_num = msg_number).first()
        sender = Drive()
        print ("Fonud a driver")
    elif len(student_numbers) > 0:
        x = Students.query.filter_by(c_num = msg_number).first()
        sender = Student()
        print ("Found a student")
    elif len(admin_numbers) > 0:
        x = Users.query.filter_by(c_num = msg_number).first()
        sender = Admin()
        print ("It's an admin...great")


    # For drivers
    if "New trip" in message_body and type(sender) is Drive:

        # New trip
        loc = message_body.split("to ",1)[1]
        print('{} created a new trip to {}'.format(x.first_name, loc))
        r_pin = randint(1000,9999)
        new_trip = Trips(d_id=x.id,
                        dest=loc,
                        pin = r_pin,
                        driver = x.first_name
                        )
        db.session.add(new_trip)
        db.session.commit()
        resp_update.message('Hi, {}. Your trip number is {} for your trip to {}. Be safe and have fun!'.format(x.first_name, r_pin, loc))

        # Find the trip in the db
        trip = Trips.query.filter_by(pin = r_pin).first()
        session['trip_pin'] = r_pin
        trip.trip_info()

        return str(resp_update)

    elif "New trip" in message_body and type(sender) is not Drive:
        print("This person isn't a driver!")
        resp_update.message("You don't have permission to create a trip.")
        return str(resp_update)

    elif "Info" in message_body and type(sender) is Drive:

        # Send a student's info to the driver
        student = message_body.split(": ", 1)[1]
        s_search = Passengers.query.filter_by(s_first = student, trip_pin = session['trip_pin']).all()
        for each in s_search:
            resp_update.message(each.all_info())
            return str(resp_update)

    elif "Info: all" in message_body and type(sender) is Drive:

        # Something went down and we need everybody's info...
        s_search = Passengers.query.filter_by(trip_pin = session['trip_pin']).all()
        for each in s_search:
            resp_update.message(each.all_info())
            return str(resp_update)

    # For students
    elif "On" in message_body and type(sender) is Student:
        trip_pin = message_body.split("On ", 1)[1]
        print("{} is on trip {}".format(x.first_name, trip_pin))
        new_passenger = Passengers(s_id = x.id,
                                   trip_pin = trip_pin,
                                   s_first = x.first_name,
                                   s_last = x.last_name,
                                   s_m_name = x.m_name,
                                   s_f_name = x.f_name,
                                   s_m_num = x.m_num,
                                   s_f_num = x.f_num)
        db.session.add(new_passenger)
        db.session.commit()
        trip = Trips.query.filter_by(pin = trip_pin).first()


        resp_update.message("Hi, {} -- You're all set! Have fun on your trip to {}.".format(x.first_name, trip.dest))
        return str(resp_update)

    # Condition for not riding back TODO: new argument (boolean) for return trip
    elif "Off" in message_body and type(sender) is Student:
        trip_pin = message_body.split("Off ", 1)[1]
        print("{} is not riding back with the group".format(x.first_name))

        passenger = Passengers.query.filter_by(s_id=x.id).filter_by(trip_pin = trip_pin).first()
        passenger.return_trip = 0
        db.session.commit()

        resp_update.message("Got it, {}. You're not riding back on the cheese.".format(x.first_name))
        return str(resp_update)

# For joining
    elif "JOIN" in message_body:

        # Question 1
        new_student = Students(c_num = msg_number)
        db.session.add(new_student)
        db.session.commit()
        resp_update.message(questions["1"])
        session['question_number'] = 1
        return str(resp_update)

    elif session['question_number'] == 1:
        # Question 2
        stud_reg = Students.query.filter_by(c_num = msg_number).first()
        stud_reg.first_name = message_body.split(" ",1)[0]
        stud_reg.last_name = message_body.split(" ",1)[1]
        db.session.commit()
        resp_update.message(questions["2"])
        session['question_number'] += 1
        return str(resp_update)

    elif session['question_number'] == 2:
        # Question 3
        stud_reg = Students.query.filter_by(c_num = msg_number).first()
        stud_reg.m_name = message_body
        db.session.commit()
        resp_update.message(questions["3"])
        session['question_number'] += 1
        return str(resp_update)

    elif session['question_number'] == 3:
        # Question 4
        stud_reg = Students.query.filter_by(c_num = msg_number).first()
        stud_reg.m_num = "+1" + str(message_body)
        db.session.commit()
        resp_update.message(questions["4"])
        session['question_number'] += 1
        return str(resp_update)

    elif session['question_number'] == 4:
        # Question 4
        stud_reg = Students.query.filter_by(c_num = msg_number).first()
        stud_reg.f_name = message_body
        db.session.commit()
        resp_update.message(questions["5"])
        session['question_number'] += 1
        return str(resp_update)

    elif session['question_number'] == 5:
        # Question 5
        stud_reg = Students.query.filter_by(c_num = msg_number).first()
        stud_reg.f_num = "+1" + str(message_body)
        db.session.commit()
        resp_update.message("Thanks for signing up, {}! You can join your first ride by sending a text to me that says 'On [PIN]' For [PIN] type the four-digit code your teacher has given you. Be safe and have fun!".format(stud_reg.first_name))
        session.pop('question_number')
        return str(resp_update)

app.run(port=80)
